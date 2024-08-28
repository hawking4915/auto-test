#coding=utf-8

from xlrd import open_workbook
from method.out_log import logger
from json import dumps,loads
from config.ProjectInfo import excel_path
from xlutils.copy import copy
from openpyxl import load_workbook,Workbook

#xlrd和xlwt处理的是xls文件，单个sheet最大行数是65535，如果有更大需要的，建议使用openpyxl函数，最大行数达到1048576。
#如果数据量超过65535就会遇到：ValueError: row index was 65536, not allowed by .xls format

#xlwings：读写，支持 xls、xlsx 格式，需要安装有 excel 软件，依赖于 pywin32
#openpyxl：读写，仅支持 xlsx 格式，不需要安装 excel 软件（相比xlrd读文件耗时慢）
#xlrd：读，支持 xls
#xlwt：写，支持 xls


# -----------------------------------------------------------
#     author：wangshinan
#     description：读取excel表格中某一sheet的全部数据——xlrd
#     param：1、data_path  excle表格路径,单独使用需传入绝对路径
#            2、sheetname  excle表格内sheet名
#     remarks：输出格式：AllData = [[第一行数据]，[第二行数据],[第三行数据]]
# -----------------------------------------------------------

def read_excel(data_path, sheetname):
    # 打开excel表格
    data = open_workbook(data_path)
    # 切换到相应sheet
    table = data.sheet_by_name(sheetname)
    # 获取表格行数
    rowNum = table.nrows
    AllData = []
    if rowNum < 2:
        print("excle内数据行数小于2（无测试数据）")
    else:
        # 从第二行（数据行）开始取数据
        for i in range(1, rowNum):
            # 获取每一行的数据，返回一个列表
            list = table.row_values(i)
            # 将每行数据添加到一个列表中
            AllData.append(list)
    return AllData, rowNum

# -----------------------------------------------------------
#     author：wangshinan
#     description：处理excel中读取出来的数据AllData
#     param：1、data_path:excle表格路径,单独使用需传入绝对路径
#            2、sheetname：excle表格内sheet名
#            3、Get_data_Mode：处理第四列参数的方式，只能填API api或GUI gui
#     remarks：list输出格式：[[list1],[list1],[list1]]
#              data输出格式：{测试编号1：[[list1],[list1]]，测试编号2：[[list1],[list1]]}
#            1、list1 GUI格式：[['参数值1','参数值2'],['功能点Xpath1', '功能点Xpath2'],['期望True_Xpath1','期望True_Xpath2'],'附加','期望','测试编号' ,序号]
#            2、list1 API格式：[{'键1': '参数值1', '键2': '参数值2'}, 'url','前提URL','附加, '期望','测试编号', 序号]
# -----------------------------------------------------------

def Get_excel_data(data_path, sheetname, Get_data_Mode):
    list = []
    AllData, rowNum = read_excel(data_path, sheetname)
    for index in range(0, len(AllData)):
        datas = []
        if index <= len(AllData):
            number = int(AllData[index][0])
            testNum = AllData[index][1]

            if Get_data_Mode == 'API' or Get_data_Mode == 'api':
                url = AllData[index][2]
                fristURL = AllData[index][3]
                test_data_last = Get_API_data(AllData[index][4])
                hope = AllData[index][5]
                attach = AllData[index][6].split('\n')
                datas.append(test_data_last)
                datas.append(url)
                datas.append(fristURL)
                datas.append(attach)
                datas.append(hope)
            else:
                clickXpath = AllData[index][2].split('\n')
                hopeXpath = AllData[index][3].split('\n')
                test_data_last = Get_GUI_data(AllData[index][4])
                hope = AllData[index][5].split('\n')
                attach = AllData[index][6].split('\n')
                datas.append(test_data_last)
                datas.append(clickXpath)
                datas.append(hopeXpath)
                datas.append(attach)
                datas.append(hope)
            index += 1
            datas.append(testNum)
            datas.append(number)
            list.append(datas)
    data = {}
    for i in list:
        if i[-2] in data:
            data[i[-2]] = data[i[-2]] + [i]
        else:
            data[i[-2]] = [i]
    return data

# -----------------------------------------------------------
#     author：wangshinan
#     description：API_处理excel中参数数据
#     param：1、TestData  excle第四列数据
#     remarks：处理之前：'username=admin\npassword=123456'
#              处理之后：{'username': 'admin', 'password': '123456'}
# -----------------------------------------------------------

def Get_API_data(TestData):
    list = TestData.split('\n')
    # print(list)
    test = {}
    for i in list:
        if ':' in i:
            td = i.split(':')
            test[td[0]] = td[1]
        else:
            td = ''
            test = td
    return test


# -----------------------------------------------------------
#     author：wangshinan
#     description：GUI_处理excel中参数数据
#     param：1、TestData  excle第四列数据
#     remarks：处理之前：'username=admin\npassword=123456'
#              处理之后：['admin', '123456']
# -----------------------------------------------------------

def Get_GUI_data(TestData):
    # 处理行数据中AllData[index][3]的参数数据
    list = TestData.split('\n')
    test = []
    for i in list:
        td = i.split('=')[1]
        test.append(td)
    return test

# -----------------------------------------------------------
#     author：wangshinan
#     description：SwaggerAPI_将响应数据写入Excel
#     param：1、text  写入的内容
#           2、excel_name 文件名
#     remarks：
# -----------------------------------------------------------

def write_excel(excel_name,sheet_name,content):
    tableTitle = ['序号', '测试编号', '接口地址', 'FirstURL', '请求参数', '期望结果', '附加', '请求方式', '功能菜单', '接口名称']
    try:
        wb = load_workbook(excel_path + excel_name)
    except:
        wb = Workbook()
        wb.save(excel_path + excel_name)
    sheet = wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    ws.append(tableTitle)
    for r in range(len(content)):
        for c in range(len(content[0])):
            sheet.cell(r + 2, c + 1).value = content[r][c]
    wb.save(excel_path + excel_name)
    wb.close()

# -----------------------------------------------------------
#     author：wangshinan
#     description：API_处理测试数据中的附加数据_将字符串转换为字典
#     param：1、attach 文件：API_TestData.xlsx中附加那一列
#     remarks：字符串1格式：'{"a":1,"b":2}'   字符串2格式：'{"a":[{"c":1}],"b":2}'
#     将字符串1转换为字典json.loads/ast.literal_eval，将字符串2转换为字典json.loads
#     ast.literal_eval局限性：无法转换字符串2
#     json.loads局限性：数组或对象之中的字符串必须使用双引号，不能使用单引号
# -----------------------------------------------------------

def processing_attachData_Dict(attach):
    list=[]
    for i in attach:
        list.append(loads(i))
    return list

# -----------------------------------------------------------
#     author：wangshinan
#     description：API_处理测试数据中的附加数据_将字典转换为json字符串
#     param：1、attach 文件：API_TestData.xlsx中附加那一列
#     remarks：Python数据类型包含所有Python基本数据类型，列表，元组，字典，自己定义的类等
#     json.dumps将一个Python数据类型转换为JSON
#     json.loads将一个JSON编码的字符串转换回一个Python数据类型
#     json.dump() 和 json.load() 来编码和解码JSON数据,主要用来读写json文件函数。
# -----------------------------------------------------------

def processing_attachData_json(attach):
    return dumps(attach)

# -----------------------------------------------------------
#     author：wangshinan
#     description：判断数据是否存在list1列表，获取下标再找list2的数据
#     param：
#     remarks：
# -----------------------------------------------------------

def IfExistList(list1,list2,data1,mark):
    try:
        index = list1.index(data1)
        data2 = list2[index]
    except Exception as e:
        logger.error("\033[1;35m%s(%s)-不存在.\033[0m" % (mark,data1))
        return None
    else:
        return data2

